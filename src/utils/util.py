import json
import os
import loguru
import openpyxl

from ..student import read_student_data


def check_file() -> bool:
    flag = True

    if not os.path.exists("data"):
        os.makedirs("data")
        flag = False
    if not os.path.exists("config"):
        os.makedirs("config")
        flag = False

    file_path = "config/config.json"
    excel_data_path = "config/data.xlsx"

    try:
        # 检查文件是否存在
        if not os.path.isfile(file_path):
            with open(file_path, "w") as config_json:
                config_dict = {"admin_name": "", "password": "", "score_min": 1, "score_max": 100}
                json.dump(config_dict, config_json, indent=4)
            flag = False

        # 检查Excel文件是否存在，如果不存在则创建
        if not os.path.isfile(excel_data_path):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet['A1'] = "学号"
            sheet['B1'] = "姓名"
            wb.save(excel_data_path)
            wb.close()
            flag = False

    except Exception as e:
        # 捕获异常并打印错误信息
        loguru.logger.error(f"文件初始化失败: {str(e)}")
        flag = False

    return flag


def find_max_min() -> tuple:
    with open('config/config.json', 'r') as config_file:
        config_data = json.load(config_file)
        min_score = config_data['score_min']
        max_score = config_data['score_max']
    return min_score, max_score


def calculate_data():
    students = read_student_data()
    student_num = len(students)

    if not os.path.exists("FinalScore"):
        os.mkdir("FinalScore")

    student_scores = {
        "思想道德素质分": {},
        "身心素质分": {},
        "审美与人文素养分": {},
        "劳动素养分": {}
    }

    # 遍历每个学生
    for stu_id, stu_name in students.items():
        # 存储其他学生对目标学生的分数
        moral_scores = []
        heart_scores = []
        human_scores = []
        labor_scores = []
        for stu in students.keys():
            if stu == stu_id:
                continue  # 跳过自己
            file_name = f"{stu}.xlsx"
            file_path = os.path.join("data/", file_name)

            if os.path.isfile(file_path):
                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    student_id, student_name, moral_score, heart_score, human_score, labor_score = row

                    if student_id == stu_id:
                        # 如果当前行的学生是目标学生，则获取其他学生对他的打分
                        moral_scores.append(int(moral_score))
                        heart_scores.append(int(heart_score))
                        human_scores.append(int(human_score))
                        labor_scores.append(int(labor_score))

        moral_scores.sort()
        heart_scores.sort()
        human_scores.sort()
        labor_scores.sort()

        # 计算分数（去掉最高分和最低分）
        final_moral_score = sum(moral_scores[1:-1]) / max(1, student_num - 2 - 1)
        final_heart_score = sum(heart_scores[1:-1]) / max(1, student_num - 2 - 1)
        final_human_score = sum(human_scores[1:-1]) / max(1, student_num - 2 - 1)
        final_labor_score = sum(labor_scores[1:-1]) / max(1, student_num - 2 - 1)

        # 保存最终分数
        student_scores["思想道德素质分"][stu_id] = final_moral_score
        student_scores["身心素质分"][stu_id] = final_heart_score
        student_scores["审美与人文素养分"][stu_id] = final_human_score
        student_scores["劳动素养分"][stu_id] = final_labor_score

    return student_scores


def create_xlsx(student_scores, score_type):
    student = read_student_data()

    if score_type == '思想道德素质分':
        score_list = sorted(student_scores[score_type].items(), key=lambda kv: (kv[1], kv[0]))
    elif score_type == '身心素质分':
        score_list = sorted(student_scores[score_type].items(), key=lambda kv: (kv[1], kv[0]))
    elif score_type == '审美与人文素养分':
        score_list = sorted(student_scores[score_type].items(), key=lambda kv: (kv[1], kv[0]))
    elif score_type == '劳动素养分':
        score_list = sorted(student_scores[score_type].items(), key=lambda kv: (kv[1], kv[0]))
    else:
        raise ValueError("Invalid score_type")

    score_list.reverse()

    # 分数区间计算
    total_students = len(score_list)
    a_percentage = round(total_students * 0.15)
    b_percentage = round(total_students * 0.35)
    c_percentage = round(total_students * 0.35)

    # 创建Excel文件
    score_headers = ["学号", "姓名", "分数", "等级", "实际分数"]
    file_name = os.path.join("FinalScore", f"{score_type}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(score_headers)

    for student_info in score_list:
        student_id, score = student_info
        student_name = student.get(student_id)
        student_score = score

        # 根据分数计算等级和分数
        if student_score >= score_list[a_percentage - 1][1]:
            level = 'A'
            score = 93
        elif student_score >= score_list[a_percentage + b_percentage - 1][1]:
            level = 'B'
            score = 91
        elif student_score >= score_list[a_percentage + b_percentage + c_percentage - 1][1]:
            level = 'C'
            score = 89
        else:
            level = 'D'
            score = 87

        ws.append([student_id, student_name, score, level, student_score])

    wb.save(file_name)
